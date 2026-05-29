import io
import re
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.page import PageMargins


# =========================
# 基本設定
# =========================
st.set_page_config(
    page_title="向精神薬チェックツール",
    page_icon="📦",
    layout="wide",
)

APP_TITLE = "向精神薬チェックツール"
OUTPUT_SHEET_NAME = "向精神クエリ"
MASTER_CSV_PATH = Path("masters/product_master.csv")
GUIDE_DIR = Path("guide")

OUTPUT_COLUMNS = [
    "商品コード",
    "商品名",
    "包装単位",
    "商品マスタ.記号",
    "数",
    "備考",
]

# 今回の対象
DEFAULT_TARGET_KEYWORDS = [
    "向精神",
    "生活改善薬",
    "毒薬",
]



STEP1_GUIDE_TEXTS_BY_FILE = {
    "step1_01.png": "「現在在庫検索」を開きます。",
    "step1_02.png": "該当支店を選択し、CSV出力",
    "step1_03.png": "ダウンロードが完了すると右上に表示されますので左クリック",
    "step1_04.png": "フォルダに入っていることを確認。<br>通常はダウンロードフォルダに入ってます",
}

GUIDE_TEXTS = {
    "step2": [
        "ダウンロードした在庫データをアップロードします。",
        "内容を確認して、資料を作成します。",
        "完成したExcelファイルをダウンロードします。",
    ],
}


# =========================
# デザイン
# =========================
st.markdown(
    """
<style>
.main .block-container {
    padding-top: 2rem;
    padding-bottom: 3rem;
    max-width: 1180px;
}
.step-card {
    background: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 18px;
    padding: 20px 22px;
    margin: 14px 0 10px 0;
    box-shadow: 0 6px 18px rgba(15, 23, 42, 0.06);
}
.step-title {
    font-size: 1.15rem;
    font-weight: 800;
    color: #0f172a;
    margin-bottom: 8px;
}
.step-note {
    color: #334155;
    line-height: 1.7;
}
.small-note {
    color: #334155;
    font-size: 0.92rem;
}
.ok-box {
    background: #ecfdf5;
    border: 1px solid #a7f3d0;
    color: #065f46;
    border-radius: 14px;
    padding: 12px 14px;
    margin: 10px 0;
}
.warn-box {
    background: #fffbeb;
    border: 1px solid #fde68a;
    color: #92400e;
    border-radius: 14px;
    padding: 12px 14px;
    margin: 10px 0;
}
.error-box {
    background: #fef2f2;
    border: 1px solid #fecaca;
    color: #991b1b;
    border-radius: 14px;
    padding: 12px 14px;
    margin: 10px 0;
}
.admin-box {
    background: #f8fafc;
    border: 1px dashed #94a3b8;
    color: #334155;
    border-radius: 14px;
    padding: 12px 14px;
    margin: 10px 0;
}
.guide-slide-count {
    text-align: center;
    font-weight: 700;
    color: #334155 !important;
    line-height: 1.9rem;
}
.guide-caption {
    text-align: center;
    color: #1f2937 !important;
    margin-top: 4px;
    margin-bottom: 8px;
    line-height: 1.6;
    font-weight: 700;
}
.guide-next-arrow {
    margin-top: 120px;
    text-align: center;
    color: #ef4444;
    font-size: 9rem;
    line-height: 0.8;
    font-weight: 900;
    text-shadow: 0 5px 14px rgba(239, 68, 68, 0.35);
}
.guide-next-arrow-label {
    margin-bottom: 10px;
    text-align: center;
    color: #b91c1c;
    font-size: 1.45rem;
    font-weight: 900;
    white-space: nowrap;
}
.upload-guide-card {
    background: #dbeafe;
    border: 2px solid #60a5fa;
    border-radius: 18px;
    padding: 20px 22px;
    margin: 12px 0 14px 0;
    width: 100%;
    box-sizing: border-box;
}
.upload-guide-title {
    font-weight: 800;
    font-size: 1.15rem;
    color: #0f172a;
    margin-bottom: 8px;
}
.upload-guide-note {
    color: #334155;
    font-size: 1rem;
    line-height: 1.8;
}
.upload-guide-em {
    font-weight: 800;
    color: #1d4ed8;
}
div[data-testid="stFileUploader"] {
    margin-top: 8px;
}
div[data-testid="stFileUploader"] section[data-testid="stFileUploaderDropzone"] {
    min-height: 120px !important;
    padding: 28px 30px !important;
    background: #eef6ff !important;
    border: 2px dashed #3b82f6 !important;
    border-radius: 16px !important;
    box-shadow: 0 0 0 1px rgba(59, 130, 246, 0.18) !important;
}
div[data-testid="stFileUploader"] section[data-testid="stFileUploaderDropzone"] * {
    color: #0f172a !important;
}
div[data-testid="stFileUploader"] section[data-testid="stFileUploaderDropzone"] button[data-testid="baseButton-secondary"] {
    background: #2563eb !important;
    color: #ffffff !important;
    border: 1px solid #1d4ed8 !important;
    border-radius: 12px !important;
    padding: 12px 22px !important;
    min-height: 50px !important;
    font-weight: 800 !important;
    font-size: 1.05rem !important;
    box-shadow: 0 4px 12px rgba(37, 99, 235, 0.35) !important;
}
div[data-testid="stFileUploader"] section[data-testid="stFileUploaderDropzone"] button[data-testid="baseButton-secondary"] * {
    color: #ffffff !important;
    font-weight: 800 !important;
}
div[data-testid="stFileUploader"] section[data-testid="stFileUploaderDropzone"] button[data-testid="baseButton-secondary"]:hover {
    background: #1d4ed8 !important;
    border-color: #1e40af !important;
}
</style>
""",
    unsafe_allow_html=True,
)


# =========================
# 共通関数
# =========================
def read_csv_safely(file_or_path) -> pd.DataFrame:
    """CSVを文字コード違いに強く読み込む。Excel由来のCP932にも対応。"""
    last_error = None
    for enc in ["utf-8-sig", "utf-8", "cp932", "shift_jis"]:
        try:
            if hasattr(file_or_path, "seek"):
                file_or_path.seek(0)
            return pd.read_csv(file_or_path, dtype=str, encoding=enc)
        except Exception as e:
            last_error = e
    raise ValueError(f"CSVを読み込めませんでした。文字コードまたは形式を確認してください。詳細: {last_error}")


def normalize_code(value) -> str:
    """商品コードの比較用。Excelで数値扱いされても文字列として揃える。"""
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def find_column(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """候補名から実在する列名を探す。"""
    normalized_map = {str(col).strip(): col for col in df.columns}
    for name in candidates:
        if name in normalized_map:
            return normalized_map[name]
    return None


def read_table(uploaded_file) -> pd.DataFrame:
    """CSV/Excelを読み込む。Excelは先頭シートを読む。"""
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return read_csv_safely(uploaded_file)
    return pd.read_excel(uploaded_file, sheet_name=0, dtype=str)


def normalize_master(master: pd.DataFrame) -> pd.DataFrame:
    """商品マスタを 商品コード・記号 の2列に整える。"""
    master = master.copy()
    master.columns = [str(c).strip() for c in master.columns]

    code_col = find_column(master, ["商品コード", "商品CD", "商品cd", "品目コード"])
    mark_col = find_column(master, ["記号", "商品マスタ.記号", "区分", "分類"])

    if code_col is None or mark_col is None:
        raise ValueError(
            "商品マスタに必要な列がありません。必要な列は「商品コード」と「記号」です。"
        )

    result = master[[code_col, mark_col]].copy()
    result.columns = ["商品コード", "記号"]

    result["商品コード"] = result["商品コード"].map(normalize_code)
    result["記号"] = result["記号"].fillna("").astype(str).str.strip()

    result = result[result["商品コード"] != ""]
    result = result[result["記号"] != ""]
    result = result.drop_duplicates(subset=["商品コード"], keep="first").reset_index(drop=True)

    return result


def load_master(master_upload=None) -> pd.DataFrame:
    """
    商品マスタを読む。
    優先順位：
    1. 管理者用タブで今回反映したマスタ
    2. masters/product_master.csv
    """
    if master_upload is not None:
        master = read_table(master_upload)
        return normalize_master(master)

    if "active_master_df" in st.session_state:
        return st.session_state["active_master_df"].copy()

    if MASTER_CSV_PATH.exists():
        master = read_csv_safely(MASTER_CSV_PATH)
        return normalize_master(master)

    raise FileNotFoundError(
        "商品マスタが見つかりません。管理者用タブで商品マスタCSVを作成するか、masters/product_master.csv を置いてください。"
    )


def filter_master(master: pd.DataFrame, selected_marks: list[str]) -> pd.DataFrame:
    """選択された記号だけ商品マスタを絞る。"""
    if not selected_marks:
        return master.iloc[0:0].copy()
    return master[master["記号"].isin(selected_marks)].reset_index(drop=True)


def df_to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Excelで開きやすいUTF-8 BOM付きCSVにする。"""
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")


def save_master_csv(df: pd.DataFrame) -> None:
    """ローカルのmasters/product_master.csvへ保存する。"""
    MASTER_CSV_PATH.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(MASTER_CSV_PATH, index=False, encoding="utf-8-sig")


def build_output(source: pd.DataFrame, master: pd.DataFrame, target_keywords: list[str]) -> pd.DataFrame:
    """
    元データと商品マスタを照合して、出力用5列を作る。
    並び順ロジックは後から追加する前提。
    """
    source = source.copy()
    source.columns = [str(c).strip() for c in source.columns]

    code_col = find_column(source, ["商品コード", "商品CD", "商品cd", "品目コード"])
    name_col = find_column(source, ["商品名", "品名", "商品名称", "正式名称"])
    pack_col = find_column(source, ["包装単位", "包装", "規格", "包装規格"])
    qty_col = find_column(source, ["数", "数量", "在庫数", "現在庫数", "在庫数量"])

    missing = []
    if code_col is None:
        missing.append("商品コード")
    if name_col is None:
        missing.append("商品名")
    if pack_col is None:
        missing.append("包装単位")
    if qty_col is None:
        missing.append("数/数量/在庫数")

    if missing:
        raise ValueError(
            "元データに必要な列がありません："
            + "、".join(missing)
            + "。列名を確認してください。"
        )

    work = source[[code_col, name_col, pack_col, qty_col]].copy()
    work.columns = ["商品コード", "商品名", "包装単位", "数"]
    work["商品コード"] = work["商品コード"].map(normalize_code)

    work["数"] = pd.to_numeric(work["数"], errors="coerce").fillna(0)

    grouped = (
        work.groupby(["商品コード", "商品名", "包装単位"], as_index=False, dropna=False)["数"]
        .sum()
    )

    merged = grouped.merge(master, on="商品コード", how="left")
    merged["記号"] = merged["記号"].fillna("").astype(str).str.strip()

    if target_keywords:
        pattern = "|".join(target_keywords)
        merged = merged[merged["記号"].str.contains(pattern, na=False)]
    else:
        merged = merged[merged["記号"] != ""]

    output = merged[["商品コード", "商品名", "包装単位", "記号", "数"]].copy()
    output["備考"] = ""
    output.columns = OUTPUT_COLUMNS

    # ---- 並び替え用の内部キー（最終出力には含めない） ----
    mark_series = output["商品マスタ.記号"].fillna("").astype(str).str.strip()
    output["sort_category"] = 9
    output.loc[mark_series.str.contains("向精神", na=False), "sort_category"] = 1
    output.loc[mark_series.str.contains("生活改善薬", na=False), "sort_category"] = 2
    output.loc[mark_series.str.contains("毒薬", na=False), "sort_category"] = 3

    # 商品名からメーカー名・OD表記を除外した並び替えキー（同一成分系を近くにまとめる）
    name_series = output["商品名"].fillna("").astype(str)
    output["sort_name_base"] = (
        name_series.str.replace(r"[｢「].*?[｣」]", "", regex=True)
        .str.replace(r"(ＯＤ|OD)", "", regex=True)
        .str.replace(r"(\d+(?:\.\d+)?)\s*mg", "", regex=True, flags=re.IGNORECASE)
        .str.strip()
    )

    # 普通錠を先、OD錠を後
    output["sort_od_priority"] = 0
    output.loc[name_series.str.upper().str.contains("OD", na=False), "sort_od_priority"] = 1
    output.loc[name_series.str.contains("ＯＤ", na=False), "sort_od_priority"] = 1

    # 規格（mg）を数値順に並べる。取得できない場合は後ろへ
    strength_extracted = name_series.str.extract(r"(\d+(?:\.\d+)?)\s*mg", flags=re.IGNORECASE)[0]
    output["sort_strength"] = pd.to_numeric(strength_extracted, errors="coerce").fillna(9999)

    # 同一薬品・同一規格内で「ﾄｰﾜ」を優先
    output["sort_maker_priority"] = 1
    output.loc[name_series.str.contains("ﾄｰﾜ", na=False), "sort_maker_priority"] = 0

    # 包装単位を自然順に並べる（100T < 500T < 1000T、同数なら通常 < B < H < その他）
    pack_series = output["包装単位"].fillna("").astype(str).str.strip()
    pack_number_extracted = pack_series.str.extract(r"(\d+(?:\.\d+)?)")[0]
    output["sort_pack_number"] = pd.to_numeric(pack_number_extracted, errors="coerce").fillna(9999)

    output["sort_pack_prefix"] = 9
    output.loc[pack_series.str.match(r"^[bB]", na=False), "sort_pack_prefix"] = 1
    output.loc[pack_series.str.match(r"^[hH]", na=False), "sort_pack_prefix"] = 2
    output.loc[~pack_series.str.match(r"^[bBhH]", na=False), "sort_pack_prefix"] = 0

    output = output.sort_values(
        by=[
            "sort_category",
            "sort_name_base",
            "sort_od_priority",
            "sort_strength",
            "sort_maker_priority",
            "sort_pack_number",
            "sort_pack_prefix",
            "商品名",
            "包装単位",
            "商品コード",
        ],
        kind="stable",
    ).reset_index(drop=True)

    # 最終出力は仕様列のみ
    output = output[OUTPUT_COLUMNS].copy()

    return output


def to_excel_bytes(output_df: pd.DataFrame) -> bytes:
    """出力Excelをメモリ上に作る。"""
    buffer = io.BytesIO()

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        output_df.to_excel(writer, index=False, sheet_name=OUTPUT_SHEET_NAME, startrow=1)

        ws = writer.book[OUTPUT_SHEET_NAME]
        now = datetime.now(ZoneInfo("Asia/Tokyo"))
        weekdays = ["月", "火", "水", "木", "金", "土", "日"]
        weekday = weekdays[now.weekday()]
        timestamp_text = (
            f"作成日時：{now.year}年{now.month}月{now.day}日({weekday}) "
            f"{now.hour}時{now.minute:02d}分"
        )
        ws["A1"] = timestamp_text
        ws.merge_cells("A1:F1")

        widths = {
            "A": 15,
            "B": 36,
            "C": 12,
            "D": 16,
            "E": 8,
            "F": 18,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        thin_side = Side(style="thin", color="D9D9D9")
        thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        timestamp_fill = PatternFill("solid", fgColor="F7F7F7")
        body_font = Font(size=11, color="000000")
        header_font = Font(size=11, bold=True, color="000000")

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.font = body_font

        ws["A1"].fill = timestamp_fill
        ws["A1"].font = body_font
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 20

        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 22

        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            row[0].alignment = Alignment(horizontal="center", vertical="center")
            row[1].alignment = Alignment(horizontal="left", vertical="center")
            row[2].alignment = Alignment(horizontal="center", vertical="center")
            row[3].alignment = Alignment(horizontal="center", vertical="center")
            row[4].alignment = Alignment(horizontal="center", vertical="center")
            row[5].alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[row[0].row].height = 21

        ws.auto_filter.ref = f"A2:F{ws.max_row}"
        ws.freeze_panes = "A3"
        ws.print_title_rows = "1:2"
        ws.print_area = f"A1:F{ws.max_row}"

        ws.page_setup.orientation = "portrait"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        ws.page_margins = PageMargins(
            left=0.1,
            right=0.1,
            top=0.2,
            bottom=0.2,
            header=0.1,
            footer=0.1,
        )

        for cell in ws["E"][2:]:
            cell.number_format = "0"

    buffer.seek(0)
    return buffer.getvalue()


def show_step(title: str, body: str):
    st.markdown(
        f"""
<div class="step-card">
  <div class="step-title">{title}</div>
  <div class="step-note">{body}</div>
</div>
""",
        unsafe_allow_html=True,
    )


def get_guide_images(prefix: str) -> list[Path]:
    """guide フォルダから指定プレフィックス画像を順序付きで取得する。"""
    if not GUIDE_DIR.exists():
        return []
    return sorted(
        [p for p in GUIDE_DIR.glob(f"{prefix}_*.png") if p.is_file()],
        key=lambda p: p.name,
    )


def render_guide_slider(
    prefix: str,
    descriptions: list[str] | None = None,
    fallback_description: str = "",
    descriptions_by_filename: dict[str, str] | None = None,
) -> None:
    """前へ/次へで切り替える簡易スライドUIを表示する。"""
    images = get_guide_images(prefix)
    if not images:
        return

    state_key = f"{prefix}_slide_index"
    if state_key not in st.session_state:
        st.session_state[state_key] = 0

    total = len(images)
    current_index = st.session_state[state_key]
    current_index = max(0, min(current_index, total - 1))
    st.session_state[state_key] = current_index

    left_space, center_col, right_space = st.columns([1, 3, 1], gap="small")
    with center_col:
        nav_container = st.container()
        with nav_container:
            prev_col, count_col, next_col = st.columns([1, 1, 1], gap="small")

            with prev_col:
                if st.button(
                    "← 前の画像",
                    key=f"{prefix}_prev",
                    disabled=(current_index == 0),
                ):
                    st.session_state[state_key] = max(0, st.session_state[state_key] - 1)
                    st.rerun()

            with count_col:
                st.markdown(
                    f"<div class='guide-slide-count'>{current_index + 1} / {total}</div>",
                    unsafe_allow_html=True,
                )

            with next_col:
                if st.button(
                    "次の画像 →",
                    key=f"{prefix}_next",
                    disabled=(current_index == total - 1),
                ):
                    st.session_state[state_key] = min(total - 1, st.session_state[state_key] + 1)
                    st.rerun()

        current_image_name = images[current_index].name
        caption = ""
        if descriptions_by_filename:
            caption = descriptions_by_filename.get(current_image_name, "")
        if not caption and descriptions:
            caption = descriptions[current_index] if current_index < len(descriptions) else ""
        if not caption:
            caption = fallback_description

        if caption:
            st.markdown(
                f"<div class='guide-caption'>{caption}</div>",
                unsafe_allow_html=True,
            )

        if prefix == "step1" and current_index == total - 1:
            image_col, arrow_col = st.columns([4, 2], gap="small")
            with image_col:
                st.image(str(images[current_index]), width=720)
            with arrow_col:
                st.markdown(
                    """
                    <div class="guide-next-arrow-label">次のステップへ</div>
                    <div class="guide-next-arrow">↓</div>
                    """,
                    unsafe_allow_html=True,
                )
        else:
            st.image(str(images[current_index]), width=720)


# =========================
# 画面
# =========================
st.title(APP_TITLE)

st.caption(
    "在庫データを取り込み、対象商品のチェック用資料を作成します。"
)

tab_make, tab_admin = st.tabs(["在庫資料作成", "管理者用"])

with tab_make:
    show_step(
        "ステップ1：在庫データをダウンロード",
        "下の画像を順番に確認しながら、在庫データを保存してください。"
    )
    render_guide_slider(
        "step1",
        fallback_description="表示された内容を確認し、必要なデータを保存します。",
        descriptions_by_filename=STEP1_GUIDE_TEXTS_BY_FILE,
    )

    show_step(
        "ステップ2：在庫データをアップロード",
        "ダウンロードした在庫データを下のアップロード欄に入れてください。"
    )
    render_guide_slider("step2", GUIDE_TEXTS.get("step2", []))

    st.markdown(
        """
        <div class="upload-guide-card">
            <div class="upload-guide-title">📂 在庫データをアップロードしてください</div>
            <div class="upload-guide-note">
                ステップ1で保存したファイルを、下のアップロード欄に入れてください。<br>
                ⬇ 下の枠へ<span class="upload-guide-em">ドラッグ＆ドロップ</span>、または
                <span class="upload-guide-em">Upload ボタン</span>から選択できます。
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    source_file = st.file_uploader(
        "在庫データファイルを選択またはドラッグ＆ドロップ",
        type=["xlsx", "xls", "csv"],
        help="ExcelまたはCSVファイルをアップロードしてください。商品コード・商品名・包装単位・数/数量/在庫数が必要です。",
        key="source_file",
    )

    st.markdown(
        '<div class="small-note">アップロード後、自動で内容を読み込みます。内容を確認してから「資料を作成する」を押してください。</div>',
        unsafe_allow_html=True,
    )

    master_file = None
    target_keywords = DEFAULT_TARGET_KEYWORDS.copy()

    st.markdown("---")

    if source_file is None:
        st.info("在庫データファイルをアップロードすると、資料作成を開始できます。")
    else:
        try:
            source_df = read_table(source_file)
            master_df = load_master(master_file)

            st.markdown('<div class="ok-box">ファイルを読み込みました。</div>', unsafe_allow_html=True)

            with st.expander("読み込み内容を確認する", expanded=False):
                st.write("在庫データ")
                st.dataframe(source_df.head(20), use_container_width=True)

            if st.button("資料を作成する", type="primary", use_container_width=True):
                output_df = build_output(source_df, master_df, target_keywords)

                if output_df.empty:
                    st.markdown(
                        '<div class="warn-box">対象データが0件でした。商品コードと商品マスタの記号を確認してください。</div>',
                        unsafe_allow_html=True,
                    )
                    st.stop()

                excel_bytes = to_excel_bytes(output_df)

                st.markdown(
                    f'<div class="ok-box">資料を作成しました。対象件数：{len(output_df)}件</div>',
                    unsafe_allow_html=True,
                )

                st.dataframe(output_df, use_container_width=True, height=420)

                show_step(
                    "ステップ3：完成ファイルをダウンロード",
                    "下のボタンから完成したExcelファイルをダウンロードしてください。"
                )

                st.download_button(
                    label="完成ファイルをダウンロード",
                    data=excel_bytes,
                    file_name=f"向精神薬リスト{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

        except Exception as e:
            st.markdown(
                f'<div class="error-box">エラー：{e}</div>',
                unsafe_allow_html=True,
            )

with tab_admin:
    show_step(
        "管理者用：商品マスタ更新",
        "商品マスタが更新された時だけ使う画面です。"
        "<br>整形前の商品一覧Excelをアップロードすると、対象の記号だけを抽出して、アプリ用の product_master.csv 形式に整えます。"
        "<br>作成したマスタで、アプリ内の masters/product_master.csv を上書きできます。"
    )

    st.markdown(
        '<div class="warn-box">通常作業者はこのタブを触らなくてOKです。商品マスタ更新時だけ使用してください。</div>',
        unsafe_allow_html=True,
    )

    raw_master_file = st.file_uploader(
        "整形前の商品一覧Excel/CSVをアップロード",
        type=["xlsx", "xls", "csv"],
        key="raw_master_file",
    )

    if raw_master_file is None:
        st.info("整形前の商品一覧ファイルをアップロードしてください。")
    else:
        try:
            raw_master_df = read_table(raw_master_file)
            normalized_master = normalize_master(raw_master_df)

            available_marks = sorted([x for x in normalized_master["記号"].dropna().unique().tolist() if str(x).strip() != ""])

            default_selected = [
                mark for mark in available_marks
                if ("向精神" in mark) or (mark == "生活改善薬") or (mark == "毒薬")
            ]

            selected_marks = st.multiselect(
                "抽出する記号を選んでください",
                options=available_marks,
                default=default_selected,
            )

            product_master = filter_master(normalized_master, selected_marks)

            c1, c2, c3 = st.columns(3)
            c1.metric("整形前行数", f"{len(raw_master_df):,}")
            c2.metric("記号あり件数", f"{len(normalized_master):,}")
            c3.metric("抽出後件数", f"{len(product_master):,}")

            with st.expander("記号別件数を見る", expanded=True):
                count_df = (
                    normalized_master["記号"]
                    .value_counts()
                    .rename_axis("記号")
                    .reset_index(name="件数")
                )
                st.dataframe(count_df, use_container_width=True, height=260)

            st.write("抽出後プレビュー")
            st.dataframe(product_master.head(100), use_container_width=True, height=360)

            csv_bytes = df_to_csv_bytes(product_master)

            col_a, col_b = st.columns(2)

            with col_a:
                st.download_button(
                    label="product_master.csv をダウンロード",
                    data=csv_bytes,
                    file_name="product_master.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

            with col_b:
                if st.button("このマスタで更新する", type="primary", use_container_width=True):
                    save_master_csv(product_master)
                    st.session_state["active_master_df"] = product_master.copy()
                    st.success("アプリ内の masters/product_master.csv を上書きしました。在庫資料作成タブでこのマスタを使用できます。")

            st.markdown(
                '<div class="small-note">「このマスタで更新する」を押すと、アプリ内の masters/product_master.csv を上書きします。更新後は在庫資料作成タブでそのまま使えます。</div>',
                unsafe_allow_html=True,
            )

        except Exception as e:
            st.markdown(
                f'<div class="error-box">エラー：{e}</div>',
                unsafe_allow_html=True,
            )
